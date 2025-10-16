import re
import collections
import math
import random
from openpyxl import Workbook

ALPHABET = list("абвгдежзийклмнопрстуфхцчшщьыэюя_")
M = len(ALPHABET)

def preprocess(text: str) -> str:
    text = text.lower().translate(str.maketrans({"ё": "е", "ъ": "ь", " ": "_"}))
    text = re.sub(r"[^а-я_]", "_", text)
    return re.sub(r"_+", "_", text).strip("_")

def letter_frequencies(text: str) -> dict:
    total = len(text)
    counter = collections.Counter(text)
    return dict(sorted(((ch, counter.get(ch, 0) / total) for ch in ALPHABET),
                       key=lambda x: x[1], reverse=True))

def bigram_frequencies(text: str, step: int = 1) -> dict:
    counter = collections.Counter(text[i:i + 2] for i in range(0, len(text) - 1, step))
    total = sum(counter.values())
    return {bg: count / total for bg, count in counter.items()}

def entropy_h(text: str, n: int) -> float:
    if not text:
        return 0
    counter = collections.Counter(text[i:i + n] for i in range(len(text) - n + 1))
    total = sum(counter.values())
    return -(1 / n) * sum((c / total) * math.log2(c / total) for c in counter.values())

def entropy_h_no_overlap(text: str, n: int) -> float:
    if not text:
        return 0
    step = n
    counter = collections.Counter(text[i:i + n] for i in range(0, len(text) - n + 1, step))
    total = sum(counter.values())
    return -(1 / n) * sum((c / total) * math.log2(c / total) for c in counter.values())

def entropy_experiments(text: str, n: int, trials: int = 50, length: int = 1000, no_overlap=False) -> float:
    results = []
    for _ in range(trials):
        start = random.randrange(max(1, len(text) - length))
        sample = text[start: start + length]
        if no_overlap:
            results.append(entropy_h_no_overlap(sample, n))
        else:
            results.append(entropy_h(sample, n))
    return sum(results) / len(results)

def save_to_excel(letter_freqs: dict, bigrams_overlap: dict, bigrams_no_overlap: dict, filename="entropy_analysis.xlsx"):
    wb = Workbook()

    ws_letters = wb.active
    ws_letters.title = "Letters"
    ws_letters.append(["Letter", "Frequency"])
    for ch, freq in letter_freqs.items():
        ws_letters.append([ch, freq])

    ws_bi_ov = wb.create_sheet("Bigrams (overlap)")
    ws_bi_ov.append(["Bigram", "Frequency"])
    for bg, freq in sorted(bigrams_overlap.items(), key=lambda item: item[1], reverse=True):
        ws_bi_ov.append([bg, freq])

    ws_bi_no = wb.create_sheet("Bigrams (no overlap)")
    ws_bi_no.append(["Bigram", "Frequency"])
    for bg, freq in sorted(bigrams_no_overlap.items(), key=lambda item: item[1], reverse=True):
        ws_bi_no.append([bg, freq])

    wb.save(filename)
    print(f"\nЧастоти збережено у файл: {filename}")

def main():
    try:
        with open("Tysyacha_i_odna_noch.txt", "r", encoding="utf-8") as f:
            text = preprocess(f.read(2 * 1024 * 1024))
    except FileNotFoundError:
        print("Файл 'Tysyacha_i_odna_noch.txt' не знайдено.")
        return

    print(f"\nКількість літер після фільтрації: {len(text)}")

    freqs_letters = letter_frequencies(text)
    print("\nТоп-30 літер:")
    for ch, freq in list(freqs_letters.items())[:30]:
        print(f"'{ch}': {freq:.5f}")

    freqs_bigrams_overlap = bigram_frequencies(text, step=1)
    freqs_bigrams_no_overlap = bigram_frequencies(text, step=2)

    print("\nТоп-10 біграм (перетин):")
    for bg, freq in sorted(freqs_bigrams_overlap.items(), key=lambda item: item[1], reverse=True)[:10]:
        print(f"'{bg}': {freq:.5f}")

    print("\nТоп-10 біграм (без перетину):")
    for bg, freq in sorted(freqs_bigrams_no_overlap.items(), key=lambda item: item[1], reverse=True)[:10]:
        print(f"'{bg}': {freq:.5f}")

    print("\nОцінка ентропії та надлишковості")
    h0 = math.log2(M)
    print(f"Максимальна ентропія H_0 = log2({M}) = {h0:.4f} біт/символ")
    print("-" * 45)

    # Ентропія з перетинами
    for n in [1, 2]:
        hn_overlap = entropy_experiments(text, n, trials=50, length=5000)
        redundancy_overlap = 1 - (hn_overlap / h0)
        print(f"Модель {n}-грам (з перетинами):")
        print(f"  Ентропія H_{n} ≈ {hn_overlap:.4f} біт/символ")
        print(f"  Надлишковість R_{n} ≈ {redundancy_overlap:.1%}")

        if n == 2:
            hn_no_overlap = entropy_experiments(text, n, trials=50, length=5000, no_overlap=True)
            redundancy_no_overlap = 1 - (hn_no_overlap / h0)
            print(f"Модель {n}-грам (без перетинів):")
            print(f"  Ентропія H_{n} ≈ {hn_no_overlap:.4f} біт/символ")
            print(f"  Надлишковість R_{n} ≈ {redundancy_no_overlap:.1%}")

    print("\nАналіз без пробілів ('_'):")
    text_no_space = text.replace("_", "")

    for n in [1, 2]:
        hn_overlap = entropy_experiments(text_no_space, n, trials=50, length=5000)
        redundancy_overlap = 1 - (hn_overlap / h0)
        print(f"Модель {n}-грам (з перетинами, без пробілів):")
        print(f"  Ентропія H_{n} ≈ {hn_overlap:.4f} біт/символ")
        print(f"  Надлишковість R_{n} ≈ {redundancy_overlap:.1%}")

        if n == 2:
            hn_no_overlap = entropy_experiments(text_no_space, n, trials=50, length=5000, no_overlap=True)
            redundancy_no_overlap = 1 - (hn_no_overlap / h0)
            print(f"Модель {n}-грам (без перетинів, без пробілів):")
            print(f"  Ентропія H_{n} ≈ {hn_no_overlap:.4f} біт/символ")
            print(f"  Надлишковість R_{n} ≈ {redundancy_no_overlap:.1%}")

    save_to_excel(freqs_letters, freqs_bigrams_overlap, freqs_bigrams_no_overlap)

if __name__ == "__main__":
    main()